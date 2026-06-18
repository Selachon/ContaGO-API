import os
import re
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

import web_io

CTX = web_io.init()
WEB = CTX.web

NOMBRE_PLANTILLA = "Subir Terceros desde Excel - Siigo Nube SF_CO.xlsm"
HOJA_TERCEROS_DIAN = "Datos de terceros"
HOJA_PLANTILLA = "Plantilla"
HOJA_PAISES = "Paises"


def normalizar_texto(texto):
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return " ".join(texto.split())


def normalizar_geo_texto(texto):
    t = normalizar_texto(texto)
    t = re.sub(r"[^a-z0-9\s]", " ", t)
    t = " ".join(t.split())

    # Normalización amplia para evitar fallos por variantes frecuentes.
    if "bogota" in t:
        return "bogota"
    if "cartagena" in t:
        return "cartagena"

    reemplazos = {
        "bogota d c": "bogota",
        "bogota dc": "bogota",
        "bogota distrito capital": "bogota",
        "bogota d.c": "bogota",
        "bogota d. c.": "bogota",
        "cartagena de indias": "cartagena",
        "distrito capital": "bogota",
    }
    for k, v in reemplazos.items():
        if t == k:
            return v
    return t


def seleccionar_archivo_reciente(carpeta):
    if not os.path.isdir(carpeta):
        return None
    archivos = [
        os.path.join(carpeta, f)
        for f in os.listdir(carpeta)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ]
    if not archivos:
        return None
    return max(archivos, key=os.path.getmtime)


def leer_terceros_desde_excel(path_excel):
    bruto = pd.read_excel(path_excel, sheet_name=HOJA_TERCEROS_DIAN, header=None)
    fila_header = None
    for i in range(min(30, len(bruto))):
        vals = [normalizar_texto(v) for v in bruto.iloc[i].tolist() if pd.notna(v)]
        unidos = " | ".join(vals)
        if "nit" in unidos and "razon social" in unidos and "ciudad" in unidos:
            fila_header = i
            break

    if fila_header is None:
        raise ValueError(f"No se encontró encabezado de '{HOJA_TERCEROS_DIAN}' en {path_excel}")

    headers = [str(x).strip() if pd.notna(x) else "" for x in bruto.iloc[fila_header].tolist()]
    df = bruto.iloc[fila_header + 1 :].copy()
    df.columns = headers
    df = df.loc[:, [c for c in df.columns if str(c).strip() != ""]]
    return df.reset_index(drop=True)


def limpiar_nit(valor):
    if pd.isna(valor):
        return ""
    s = str(valor).strip()
    if s.lower() == "nan":
        return ""
    s = s.replace(" ", "").replace("-", "")
    if s.endswith(".0"):
        s = s[:-2]
    s = "".join(ch for ch in s if ch.isdigit())
    return s


def tipo_identificacion_desde_nit(nit):
    if len(nit) == 9:
        try:
            n = int(nit)
            if 70000000 <= n <= 999000000:
                return "31"
        except ValueError:
            pass
    return "13"


def dividir_nombre_persona(razon_social):
    partes = [p for p in str(razon_social).strip().split() if p]
    if not partes:
        return "", ""
    if len(partes) == 1:
        return partes[0], ""
    if len(partes) == 2:
        return partes[0], partes[1]
    if len(partes) == 3:
        return " ".join(partes[:2]), partes[2]

    nombres_comunes = {
        "jose", "antonio", "juan", "manuel", "francisco", "luis", "javier", "miguel", "carlos", "angel",
        "jesus", "david", "daniel", "pedro", "alejandro", "maria", "alberto", "pablo", "fernando", "rafael",
        "jorge", "ramon", "sergio", "enrique", "andres", "diego", "adrian", "vicente", "victor", "alvaro",
        "ignacio", "raul", "eduardo", "ivan", "oscar", "ruben", "joaquin", "santiago", "mario", "roberto",
        "gabriel", "marcos", "alfonso", "jaime", "ricardo", "hugo", "julio", "emilio", "martin", "salvador",
        "guillermo", "mohamed", "nicolas", "tomas", "jordi", "julian", "gonzalo", "agustin", "cristian", "cesar",
        "marc", "felix", "joan", "josep", "samuel", "sebastian", "lucas", "hector", "felipe", "ismael",
        "alfredo", "domingo", "aitor", "alex", "mariano", "rodrigo", "mateo", "alexander", "iker", "marco",
        "xavier", "esteban", "arturo", "gregorio", "lorenzo", "dario", "borja", "albert", "aaron", "joel",
        "isaac", "eugenio", "cristobal", "eric", "jonathan", "christian", "mohammed", "pau", "german", "omar",
        "carmen", "ana", "isabel", "dolores", "pilar", "teresa", "rosa", "josefa", "cristina", "laura",
        "angeles", "elena", "antonia", "lucia", "marta", "francisca", "mercedes", "luisa", "concepcion", "rosario",
        "paula", "sara", "raquel", "rocio", "eva", "patricia", "beatriz", "victoria", "juana", "manuela",
        "julia", "andrea", "belen", "alba", "silvia", "esther", "irene", "nuria", "encarnacion", "montserrat",
        "sandra", "angela", "monica", "alicia", "inmaculada", "yolanda", "mar", "sonia", "marina", "sofia",
        "susana", "margarita", "claudia", "natalia", "carolina", "ines", "alejandra", "daniela", "carla", "veronica",
        "amparo", "gloria", "lourdes", "nieves", "luz", "soledad", "noelia", "lorena", "fatima", "begona",
        "blanca", "olga", "nerea", "miriam", "clara", "consuelo", "asuncion", "milagros", "esperanza", "martina",
        "lidia", "catalina", "adriana", "celia", "anna", "aurora", "magdalena", "emilia", "elisa", "vanesa",
        "ainhoa", "virginia", "eugenia", "diana", "gema", "alexandra", "valeria", "tatiana", "leonor",
    }

    def score_nombres(tokens):
        return sum(1 for x in tokens if normalizar_texto(x) in nombres_comunes)

    mitad = len(partes) // 2
    bloque_izq = partes[:mitad]
    bloque_der = partes[mitad:]

    # Si los nombres parecen venir al final (p.ej. CASTRO ABUCHAIBE TATIANA LEONOR), invierte.
    if score_nombres(bloque_der) > score_nombres(bloque_izq):
        nombres = " ".join(bloque_der)
        apellidos = " ".join(bloque_izq)
        return nombres, apellidos

    # Caso usual en Colombia: NOMBRES + APELLIDOS
    nombres = " ".join(partes[:2])
    apellidos = " ".join(partes[-2:])
    return nombres, apellidos


def limpiar_telefono(valor):
    if pd.isna(valor):
        return ""
    s = str(valor)
    # Quitar palabras/letras y caracteres especiales; conservar solo dígitos.
    s = re.sub(r"[^0-9]", "", s)

    # Si viene con indicativo país Colombia y supera 10, quitar 57 inicial.
    if len(s) > 10 and s.startswith("57"):
        s = s[2:]

    # Si aún supera 10 y trae indicativo ciudad (601, 602, ...), quitarlo.
    indicativos_ciudad = {"601", "602", "603", "604", "605", "606", "607", "608"}
    while len(s) > 10 and s[:3] in indicativos_ciudad:
        s = s[3:]

    # Limpieza adicional de prefijos 0 cuando el número es largo.
    while len(s) > 10 and s.startswith("0"):
        s = s[1:]

    # Si sigue largo, conservar los últimos 10 dígitos.
    if len(s) > 10:
        s = s[-10:]

    return s


def a_mayusculas(valor):
    if pd.isna(valor):
        return ""
    return str(valor).strip().upper()


def limpiar_correo(valor):
    if pd.isna(valor):
        return ""
    s = str(valor).strip().lower()
    if not s:
        return ""

    candidatos = re.split(r"[;,\s]+", s)
    patron = re.compile(r"^[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}$")
    for c in candidatos:
        c = c.strip()
        if patron.match(c):
            return c
    return ""


def calcular_digito_verificacion(nit):
    nit = "".join(ch for ch in str(nit) if ch.isdigit())
    if not nit:
        return ""

    # Pesos DIAN de derecha a izquierda.
    pesos = [3, 7, 13, 17, 19, 23, 29, 37, 41, 43, 47, 53, 59, 67, 71]
    nit_rev = nit[::-1]

    total = 0
    for i, ch in enumerate(nit_rev):
        if i >= len(pesos):
            return ""
        total += int(ch) * pesos[i]

    residuo = total % 11
    if residuo in (0, 1):
        return str(residuo)
    return str(11 - residuo)


def construir_diccionario_geo(path_plantilla):
    bruto = pd.read_excel(path_plantilla, sheet_name=HOJA_PAISES, header=None)
    fila_header = None
    for i in range(min(40, len(bruto))):
        vals = [normalizar_texto(v) for v in bruto.iloc[i].tolist() if pd.notna(v)]
        unidos = " | ".join(vals)
        if "estado / departamento" in unidos and "codigo ciudad" in unidos:
            fila_header = i
            break
    if fila_header is None:
        raise ValueError("No se encontró encabezado válido en hoja Paises")

    headers = [str(x).strip() if pd.notna(x) else "" for x in bruto.iloc[fila_header].tolist()]
    geo = bruto.iloc[fila_header + 1 :].copy()
    geo.columns = headers
    geo = geo.loc[:, [c for c in geo.columns if str(c).strip() != ""]]

    cols_norm = {normalizar_texto(c): c for c in geo.columns}
    c_pais = cols_norm.get("pais")
    c_dep = cols_norm.get("estado / departamento")
    c_ciudad = cols_norm.get("ciudad")
    c_cod_pais = cols_norm.get("codigo pais")
    c_cod_dep = cols_norm.get("codigo estado / departamento")
    c_cod_ciudad = cols_norm.get("codigo ciudad")

    faltantes = [k for k, v in {
        "Pais": c_pais,
        "Estado / Departamento": c_dep,
        "Ciudad": c_ciudad,
        "Codigo pais": c_cod_pais,
        "Codigo Estado / Departamento": c_cod_dep,
        "Codigo ciudad": c_cod_ciudad,
    }.items() if v is None]
    if faltantes:
        raise ValueError(f"Faltan columnas en hoja Paises: {', '.join(faltantes)}")

    geo = geo[[c_pais, c_dep, c_ciudad, c_cod_pais, c_cod_dep, c_cod_ciudad]].copy()
    geo.columns = ["pais", "dep", "ciudad", "cod_pais", "cod_dep", "cod_ciudad"]

    geo["k_pais"] = geo["pais"].map(normalizar_geo_texto)
    geo["k_dep"] = geo["dep"].map(normalizar_geo_texto)
    geo["k_ciudad"] = geo["ciudad"].map(normalizar_geo_texto)

    idx_ciudad = {}
    idx_dep = {}
    idx_ciudad_pais = {}
    for _, r in geo.iterrows():
        kc = (r["k_pais"], r["k_dep"], r["k_ciudad"])
        kd = (r["k_pais"], r["k_dep"])
        kcp = (r["k_pais"], r["k_ciudad"])
        if kc not in idx_ciudad:
            idx_ciudad[kc] = (str(r["cod_dep"]).strip(), str(r["cod_ciudad"]).strip(), str(r["cod_pais"]).strip())
        if kd not in idx_dep:
            idx_dep[kd] = (str(r["cod_dep"]).strip(), str(r["cod_pais"]).strip())
        if kcp not in idx_ciudad_pais:
            idx_ciudad_pais[kcp] = (str(r["cod_dep"]).strip(), str(r["cod_ciudad"]).strip(), str(r["cod_pais"]).strip())

    return idx_ciudad, idx_dep, idx_ciudad_pais


def mapear_terceros(df_terceros, path_plantilla):
    idx_ciudad, idx_dep, idx_ciudad_pais = construir_diccionario_geo(path_plantilla)

    cols_norm = {normalizar_texto(c): c for c in df_terceros.columns}
    c_nit = cols_norm.get("nit")
    c_razon = cols_norm.get("razon social")
    c_dir = cols_norm.get("direccion")
    c_pais = cols_norm.get("pais")
    c_dep = cols_norm.get("departamento")
    c_ciudad = cols_norm.get("ciudad")
    c_tel = cols_norm.get("telefono")
    c_mail = cols_norm.get("correo")

    if c_nit is None or c_razon is None:
        raise ValueError("La hoja Datos de terceros debe incluir columnas NIT y Razón social")

    out = []
    for _, r in df_terceros.iterrows():
        nit = limpiar_nit(r.get(c_nit, ""))
        if not nit:
            continue

        razon = a_mayusculas(r.get(c_razon, ""))
        if not razon:
            continue

        tipo_id = tipo_identificacion_desde_nit(nit)
        tipo = "Empresa" if tipo_id == "31" else "Es persona"

        nombres = ""
        apellidos = ""
        razon_social = ""
        if tipo_id == "31":
            razon_social = a_mayusculas(razon)
        else:
            nombres, apellidos = dividir_nombre_persona(razon)
            nombres = a_mayusculas(nombres)
            apellidos = a_mayusculas(apellidos)

        pais = "Colombia"
        if c_pais is not None and pd.notna(r.get(c_pais)):
            pais = str(r.get(c_pais)).strip() or "Colombia"
        dep = "" if c_dep is None or pd.isna(r.get(c_dep)) else str(r.get(c_dep)).strip()
        ciudad = "" if c_ciudad is None or pd.isna(r.get(c_ciudad)) else str(r.get(c_ciudad)).strip()

        k_pais = normalizar_geo_texto(pais)
        k_dep = normalizar_geo_texto(dep)
        k_ciudad = normalizar_geo_texto(ciudad)

        cod_pais = "COL" if k_pais == "colombia" else ""
        cod_dep = ""
        cod_ciudad = ""

        if (k_pais, k_dep, k_ciudad) in idx_ciudad:
            cod_dep, cod_ciudad, cod_pais_geo = idx_ciudad[(k_pais, k_dep, k_ciudad)]
            if cod_pais_geo:
                cod_pais = cod_pais_geo
        elif (k_pais, k_ciudad) in idx_ciudad_pais:
            cod_dep, cod_ciudad, cod_pais_geo = idx_ciudad_pais[(k_pais, k_ciudad)]
            if cod_pais_geo:
                cod_pais = cod_pais_geo
        elif (k_pais, k_dep) in idx_dep:
            cod_dep, cod_pais_geo = idx_dep[(k_pais, k_dep)]
            if cod_pais_geo:
                cod_pais = cod_pais_geo

        direccion = "" if c_dir is None or pd.isna(r.get(c_dir)) else str(r.get(c_dir)).strip()
        telefono = "" if c_tel is None else limpiar_telefono(r.get(c_tel))
        correo = "" if c_mail is None else limpiar_correo(r.get(c_mail))

        out.append({
            "Identificación (Obligatorio)": nit,
            "Dígito de verificación": calcular_digito_verificacion(nit),
            "Tipo identificación (Obligatorio)": tipo_id,
            "Tipo (Obligatorio)": tipo,
            "Razón social (Obligatorio)": razon_social,
            "Nombres del tercero (Obligatorio)": nombres,
            "Apellidos del tercero (Obligatorio)": apellidos,
            "Dirección": direccion,
            "Código país": cod_pais,
            "Código departamento/estado": cod_dep,
            "Código ciudad": cod_ciudad,
            "Teléfono principal": telefono,
            "Tipo de régimen IVA": "2" if tipo_id == "31" else "0",
            "Correo electrónico contacto principal": correo,
            "Clientes": "",
            "Estado": "Activo",
        })

    salida = pd.DataFrame(out)
    if salida.empty:
        return salida

    salida = salida.drop_duplicates(subset=["Identificación (Obligatorio)"], keep="first")
    return salida


def escribir_en_plantilla(path_plantilla, path_salida, df_salida):
    wb = load_workbook(path_plantilla, keep_vba=True)
    ws = wb[HOJA_PLANTILLA]
    wb.active = wb.sheetnames.index(HOJA_PLANTILLA)

    if HOJA_PAISES in wb.sheetnames:
        ws_paises = wb[HOJA_PAISES]
        wb.remove(ws_paises)

    # Algunos archivos de SIIGO traen encabezados con codificación irregular
    # y, en ciertos casos, el encabezado no está estrictamente en la fila 1.
    # Detectamos dinámicamente la fila de encabezado por columnas clave.
    fila_header = 1
    for r in range(1, min(25, ws.max_row) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        vals_norm = [normalizar_texto(v) for v in vals if v not in (None, "")]
        unidos = " | ".join(vals_norm)
        if "identificacion" in unidos and "tipo identificacion" in unidos and "correo electronico contacto principal" in unidos:
            fila_header = r
            break

    headers = [ws.cell(row=fila_header, column=c).value for c in range(1, ws.max_column + 1)]
    idx_col = {normalizar_texto(h): i + 1 for i, h in enumerate(headers) if h is not None and str(h).strip() != ""}

    fila_inicio_datos = fila_header + 1
    for fila in range(fila_inicio_datos, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=fila, column=col).value = None

    for i, (_, r) in enumerate(df_salida.iterrows(), start=fila_inicio_datos):
        for col_df, val in r.items():
            k = normalizar_texto(col_df)
            if k in idx_col:
                ws.cell(row=i, column=idx_col[k]).value = val

    os.makedirs(os.path.dirname(path_salida), exist_ok=True)
    wb.save(path_salida)


def main():
    plantilla = CTX.plantilla_terceros if WEB else NOMBRE_PLANTILLA
    if not plantilla or not os.path.exists(plantilla):
        if WEB:
            web_io.error("plantilla_faltante",
                         "No se encontró la plantilla de terceros para esta empresa")
        print(f"No se encontró la plantilla en la raíz: {plantilla}")
        input("Presiona Enter para cerrar...")
        return

    if WEB:
        fuentes = list(CTX.dians)
    else:
        fuentes = []
        for carpeta in [os.path.join("Ventas", "Entrada"), os.path.join("Compras", "Entrada")]:
            archivo = seleccionar_archivo_reciente(carpeta)
            if archivo is not None:
                fuentes.append(archivo)

    if not fuentes:
        if WEB:
            web_io.error("sin_entrada", "No se recibió archivo DIAN de entrada")
        print("No se encontraron archivos de entrada en Ventas/Entrada ni Compras/Entrada")
        input("Presiona Enter para cerrar...")
        return

    terceros_acum = []
    usados = []
    for f in fuentes:
        try:
            xls = pd.ExcelFile(f)
            if HOJA_TERCEROS_DIAN in xls.sheet_names:
                terceros_acum.append(leer_terceros_desde_excel(f))
                usados.append(f)
        except Exception:
            pass

    if not terceros_acum:
        if WEB:
            web_io.error("sin_hoja_terceros",
                         "No se encontró la hoja 'Datos de terceros' en el archivo DIAN")
        print("No se encontró la hoja 'Datos de terceros' en los archivos de entrada detectados.")
        input("Presiona Enter para cerrar...")
        return

    base = pd.concat(terceros_acum, ignore_index=True)
    salida_df = mapear_terceros(base, plantilla)

    if WEB:
        carpeta = CTX.out
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        carpeta = os.path.join("Terceros", "Salida", f"Proceso_generado_{timestamp}")
    archivo_salida = os.path.join(carpeta, "Subir_Terceros_Siigo.xlsm")
    escribir_en_plantilla(plantilla, archivo_salida, salida_df)

    if WEB:
        web_io.ok(
            {"registros": int(len(salida_df))},
            [{"tipo": "terceros", "path": archivo_salida}],
        )
        return

    print("=== Proceso de terceros finalizado ===")
    print("Archivos DIAN usados:")
    for u in usados:
        print(f" - {u}")
    print(f"Registros únicos generados: {len(salida_df)}")
    print(f"Archivo generado: {archivo_salida}")
    input("Presiona Enter para cerrar...")


if __name__ == "__main__":
    main()

